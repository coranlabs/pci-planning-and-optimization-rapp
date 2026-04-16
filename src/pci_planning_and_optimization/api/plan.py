# Copyright 2025-2026 coRAN LABS Private Limited
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from __future__ import annotations

import io
import logging
import threading
from typing import Any

from pci_planning_and_optimization.algorithm.coloring import run_optimization
from pci_planning_and_optimization.algorithm.conflict_graph import all_conflicts, prepare_network
from pci_planning_and_optimization.models import (
    Cell,
    NeighborRelation,
    Network,
    RelationSource,
    Technology,
)

_log = logging.getLogger(__name__)

__all__ = [
    "EXPORT_HEADERS",
    "PlanStore",
    "network_to_plan",
    "parse_plan",
    "plan_template_bytes",
    "plan_workbook_bytes",
    "sample_plans",
]

TEMPLATE_HEADERS = [
    "cell_id", "technology", "lat", "lng", "azimuth",
    "pci", "arfcn", "duplex", "cell_type", "neighbors",
]

EXPORT_HEADERS = [
    "cell_id", "technology", "lat", "lng", "azimuth",
    "pci_before", "pci", "arfcn", "duplex", "cell_type", "neighbors",
]

_WIDTHS = {"cell_id": 22, "pci": 7, "pci_before": 11, "cell_type": 10, "neighbors": 46}

_REQUIRED = ("cell_id", "technology", "lat", "lng", "pci", "neighbors")
_ALLOWED_CELL_TYPES = {"macro", "small", "das", "indoor"}

_TECHNOLOGY_ALIASES = {
    "nr": Technology.NR, "5g": Technology.NR, "5gnr": Technology.NR,
    "nr5g": Technology.NR, "5g-nr": Technology.NR, "nr-5g": Technology.NR,
    "lte": Technology.LTE, "4g": Technology.LTE, "4glte": Technology.LTE,
    "lte4g": Technology.LTE, "eutran": Technology.LTE, "e-utran": Technology.LTE,
}

_DUPLEX_ALIASES = {"fdd": "FDD", "tdd": "TDD"}

EDGE_CLASSES = ("collision", "confusion", "mod3", "mod4", "mod30")

_SEVERITY = {"collision": 0, "confusion": 1, "mod3": 2, "mod4": 3, "mod30": 4}


def plan_workbook_bytes(
    rows: list[tuple[Any, ...]], headers: list[str] | None = None,
) -> bytes:
    import openpyxl

    headers = headers or TEMPLATE_HEADERS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PCI Plan"
    ws.append(headers)
    for i, h in enumerate(headers):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = _WIDTHS.get(h, 11)
    bold = openpyxl.styles.Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}"
    for row in rows:
        ws.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def plan_template_bytes() -> bytes:
    return plan_workbook_bytes([
        ("SITE001-1", "lte", 53.3498, -6.2603, 0, 10, 1800, "FDD", "macro",
         "SITE001-2,SITE001-3,SITE002-1"),
        ("SITE001-2", "lte", 53.3498, -6.2603, 120, 11, 1800, "FDD", "macro",
         "SITE001-1,SITE001-3,SITE002-1"),
        ("SITE001-3", "lte", 53.3498, -6.2603, 240, 12, 1800, "TDD", "macro",
         "SITE001-1,SITE001-2"),
        ("SITE002-1", "lte", 53.3520, -6.2650, 0, 10, 1800, "", "macro",
         "SITE001-1,SITE001-2"),
        ("SITE003-1", "nr", 53.3540, -6.2700, 0, 700, 632628, "", "macro",
         "SITE003-2"),
        ("SITE003-2", "nr", 53.3540, -6.2700, 120, 701, 632628, "", "macro",
         "SITE003-1"),
    ])



def _norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _to_float(value: Any) -> float | None:
    try:
        return float(_norm(value))
    except (TypeError, ValueError):
        return None


def _to_int(value: Any) -> int | None:
    f = _to_float(value)
    return None if f is None else int(f)


def _split_neighbors(raw: Any) -> list[str]:
    text = _norm(raw).replace(";", ",")
    return [t.strip() for t in text.split(",") if t.strip()]


def parse_plan(raw: bytes) -> dict[str, Any]:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        headers = [
            _norm(c.value).lower() or f"col{i}" for i, c in enumerate(ws[1])
        ]
    except Exception as exc:
        return {"error": f"Failed to read workbook: {exc}"}

    missing = [h for h in _REQUIRED if h not in headers]
    if missing:
        return {
            "error": (
                f"Missing required column(s): {', '.join(missing)}. "
                f"Expected header row: {', '.join(TEMPLATE_HEADERS)}."
            )
        }

    cells: list[Cell] = []
    neighbors_by_cell: dict[str, list[str]] = {}
    invalid: list[dict[str, Any]] = []
    seen: set[str] = set()

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        rec = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

        cell_id = _norm(rec.get("cell_id"))
        if not cell_id:
            invalid.append({"row": row_num, "reason": "cell_id is empty"})
            continue
        if cell_id in seen:
            invalid.append({"row": row_num, "cell_id": cell_id,
                            "reason": "duplicate cell_id"})
            continue

        lat, lng = _to_float(rec.get("lat")), _to_float(rec.get("lng"))
        pci = _to_int(rec.get("pci"))
        if lat is None or lng is None:
            invalid.append({"row": row_num, "cell_id": cell_id,
                            "reason": "lat/lng missing or not numeric"})
            continue
        if pci is None:
            invalid.append({"row": row_num, "cell_id": cell_id,
                            "reason": "pci missing or not numeric"})
            continue

        tech_raw = _norm(rec.get("technology")).lower().replace(" ", "")
        tech = _TECHNOLOGY_ALIASES.get(tech_raw)
        if tech is None:
            invalid.append({
                "row": row_num, "cell_id": cell_id,
                "reason": (f"technology {_norm(rec.get('technology'))!r} not "
                           f"recognised — expected one of: "
                           f"{', '.join(sorted(_TECHNOLOGY_ALIASES))}"),
            })
            continue
        pci_max = 1007 if tech is Technology.NR else 503
        if not 0 <= pci <= pci_max:
            invalid.append({"row": row_num, "cell_id": cell_id,
                            "reason": f"pci {pci} outside 0..{pci_max} for {tech.value}"})
            continue

        cell_type = _norm(rec.get("cell_type")).lower() or "macro"
        if cell_type not in _ALLOWED_CELL_TYPES:
            cell_type = "macro"

        arfcn = _to_int(rec.get("arfcn")) or 1
        fields: dict[str, Any] = {
            "id": cell_id,
            "technology": tech,
            "mo_class": "NRCellDU" if tech is Technology.NR else "EUtranCellFDD",
            "pci": pci,
            "lat": lat,
            "lon": lng,
            "azimuth": _to_float(rec.get("azimuth")),
            "cell_type": cell_type,
        }
        if tech is Technology.NR:
            fields["arfcn_dl"] = arfcn
        else:
            duplex_raw = _norm(rec.get("duplex")).lower()
            if duplex_raw and duplex_raw not in _DUPLEX_ALIASES:
                invalid.append({
                    "row": row_num, "cell_id": cell_id,
                    "reason": (f"duplex {_norm(rec.get('duplex'))!r} not "
                               f"recognised — expected FDD or TDD"),
                })
                continue
            duplex = _DUPLEX_ALIASES.get(duplex_raw, "FDD")
            fields["earfcn_dl"] = arfcn
            if duplex == "FDD":
                fields["earfcn_ul"] = arfcn
            fields["duplex"] = duplex
            fields["mo_class"] = f"EUtranCell{duplex}"
            fields["pci_components"] = (pci // 3, pci % 3)

        try:
            cells.append(Cell(**fields))
        except Exception as exc:
            invalid.append({"row": row_num, "cell_id": cell_id, "reason": str(exc)})
            continue

        seen.add(cell_id)
        neighbors_by_cell[cell_id] = _split_neighbors(rec.get("neighbors"))

    if not cells:
        return {
            "error": "No usable rows found.",
            "invalid_rows": invalid[:50],
            "invalid_count": len(invalid),
        }

    known = {c.id for c in cells}
    relations: list[NeighborRelation] = []
    seen_pairs: set[tuple[str, str]] = set()
    dangling = 0
    for src, targets in neighbors_by_cell.items():
        for tgt in targets:
            if tgt not in known:
                dangling += 1
                continue
            for a, b in ((src, tgt), (tgt, src)):
                if a == b or (a, b) in seen_pairs:
                    continue
                seen_pairs.add((a, b))
                relations.append(
                    NeighborRelation(
                        source_cell_id=a,
                        target_cell_id=b,
                        relation_source=RelationSource.REAL,
                    )
                )

    _log.info(
        "plan import: %d cells, %d relations, %d invalid rows, %d dangling neighbours",
        len(cells), len(relations), len(invalid), dangling,
    )
    return {
        "network": Network(cells, relations),
        "rows": len(cells),
        "invalid_rows": invalid[:50],
        "invalid_count": len(invalid),
        "dangling_neighbors": dangling,
    }


def _edges(network: Network, config: Any) -> list[dict[str, Any]]:
    best: dict[tuple[str, str], dict[str, Any]] = {}
    for tech in (Technology.LTE, Technology.NR):
        bundle = all_conflicts(network, tech)
        for cls, edges in (("collision", bundle.collisions),
                           ("confusion", bundle.confusions),
                           ("mod3", bundle.mod3),
                           ("mod4", bundle.mod4),
                           ("mod30", bundle.mod30)):
            for e in edges:
                key = (e.cell_a_id, e.cell_b_id)
                prev = best.get(key)
                if prev is not None and _SEVERITY[prev["type"]] <= _SEVERITY[cls]:
                    continue
                best[key] = {
                    "a": e.cell_a_id,
                    "b": e.cell_b_id,
                    "type": cls,
                    "pci_a": e.pci_a,
                    "pci_b": e.pci_b,
                }
    return sorted(best.values(), key=lambda d: _SEVERITY[d["type"]])


def _cell_rows(network: Network) -> list[dict[str, Any]]:
    return [
        {
            "id": c.id,
            "pci": c.pci,
            "lat": c.lat,
            "lng": c.lon,
            "azimuth": c.azimuth,
            "tech": c.technology.value,
            "cell_type": c.cell_type,
        }
        for c in network.cells.values()
    ]


def _counts(edges: list[dict[str, Any]]) -> dict[str, int]:
    out = {cls: sum(1 for e in edges if e["type"] == cls) for cls in EDGE_CLASSES}
    out["total"] = len(edges)
    out["hard"] = out["collision"] + out["confusion"]
    return out


def _planning_config(config: Any, n_cells: int) -> Any:
    cfg = config.model_copy(deep=True)
    cfg.convergence.per_run_budget_pct = 1.0
    cfg.convergence.per_pass_budget_pct = 0.5
    cfg.convergence.max_absolute_changes = max(n_cells, 1)
    return cfg


def _export_rows(
    network: Network,
    neighbors: dict[str, list[str]],
    pci_before: dict[str, int],
) -> list[tuple[Any, ...]]:
    return [
        (
            c.id,
            c.technology.value,
            c.lat,
            c.lon,
            c.azimuth,
            pci_before.get(c.id) if pci_before.get(c.id) != c.pci else "-",
            c.pci,
            c.arfcn_dl if c.technology is Technology.NR else c.earfcn_dl,
            c.duplex or "",
            c.cell_type,
            ",".join(neighbors.get(c.id, ())),
        )
        for c in network.cells.values()
    ]


