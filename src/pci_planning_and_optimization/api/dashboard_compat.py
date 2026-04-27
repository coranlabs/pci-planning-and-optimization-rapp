# Copyright 2025-2026 coRAN LABS Private Limited
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from __future__ import annotations

import hashlib
import io
import json
import logging
import threading
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Any


def _utcnow_iso() -> str:
    return datetime.now(UTC).isoformat()


_COMMIT_CONFIRM_WINDOW_S = 15 * 60

_APPLIED_MARKER = "_pci_overrides_applied"

_UNASSIGNED_REGION = "Unassigned"

_HISTORY_STEP_S = 60
_HISTORY_MAX = 24 * 60


_FULL_NAMESPACE = {"LTE": (0, 504), "NR": (0, 1008)}


def _pci_pool_range(config: Any, cell: Any) -> tuple[int, int]:
    tech = cell.technology.name
    default = _FULL_NAMESPACE.get(tech, (0, 504))
    try:
        pools = getattr(config.pools, tech.lower())
        rng = getattr(pools, cell.cell_type, None)
    except AttributeError:
        return default
    if not rng or len(rng) != 2:
        return default
    return int(rng[0]), int(rng[1])


def _change_record(cell: Any, new_pci: int) -> dict[str, Any]:
    record: dict[str, Any] = {
        "cell_id": cell.dn or cell.id,
        "technology": cell.technology.name.lower(),
        "mo_class": cell.mo_class,
        "pci_new": new_pci,
    }
    if cell.technology.name == "LTE":
        record["pci_components_new"] = {"group": new_pci // 3, "sub": new_pci % 3}
    return record


def compute_pulse(
    cells: list[dict[str, Any]],
    conflicts: list[dict[str, Any]],
) -> dict[str, Any]:
    total_cells = len(cells)
    active_cells = sum(1 for c in cells if c.get("status") == "active")

    affected_ues = sum(int(c.get("affectedUe") or 0) for c in conflicts)

    region_by_cell: dict[str, str] = {
        c["id"]: c.get("region", "—") for c in cells if c.get("id")
    }
    region_conflicts: dict[str, int] = {}
    region_ues: dict[str, int] = {}
    for cf in conflicts:
        regions_touched = {
            region_by_cell.get(cid)
            for cid in (cf.get("cells") or [])
            if region_by_cell.get(cid)
        }
        ue = int(cf.get("affectedUe") or 0)
        for r in regions_touched:
            region_conflicts[r] = region_conflicts.get(r, 0) + 1
            region_ues[r] = region_ues.get(r, 0) + ue
    worst_region: dict[str, Any] | None = None
    if region_conflicts:
        name = max(region_conflicts, key=lambda r: region_conflicts[r])
        worst_region = {
            "name": name,
            "conflicts": region_conflicts[name],
            "ues": region_ues.get(name, 0),
        }

    mod3_cell_ids = {
        cid
        for cf in conflicts
        if cf.get("type") == "mod3"
        for cid in (cf.get("cells") or [])
    }

    return {
        "activeCells": active_cells,
        "totalCells": total_cells,
        "affectedUes": affected_ues,
        "conflictCount": len(conflicts),
        "worstRegion": worst_region,
        "mod3Cells": len(mod3_cell_ids),
    }


_log = logging.getLogger(__name__)

_THP_STEP_S = 60
_THP_MAX_AGE_S = 7 * 86400
_TS_RANGES: dict[str, int] = {
    "15m": 15 * 60, "1h": 3600, "6h": 6 * 3600, "24h": 86400, "7d": 7 * 86400,
}
_TS_DEFAULT = "1h"
_TS_TARGET_POINTS = 120


def _resample(pts: list[list[float]], span: float) -> list[dict[str, Any]]:
    if not pts:
        return []
    now = time.time()
    step = max(_THP_STEP_S, span / _TS_TARGET_POINTS)
    slots = int(span / step)
    start = now - slots * step

    acc = [[0.0, 0.0, 0] for _ in range(slots)]
    for ts, dl, ul in pts:
        i = int((ts - start) / step)
        if 0 <= i < slots:
            acc[i][0] += dl
            acc[i][1] += ul
            acc[i][2] += 1

    filled = [i for i, a in enumerate(acc) if a[2]]
    if not filled:
        return []

    out: list[dict[str, Any]] = []
    last = (0.0, 0.0)
    for i in range(filled[0], filled[-1] + 1):
        dl, ul, n = acc[i]
        if n:
            last = (dl / n, ul / n)
        stamp = datetime.fromtimestamp(start + (i + 1) * step, UTC)
        out.append({
            "t": stamp.strftime("%H:%M"), "iso": stamp.isoformat(),
            "dl": round(last[0]), "ul": round(last[1]),
        })
    return out


class ThroughputHistory:

    def __init__(self, path: str = "") -> None:
        self._path = Path(path) if path else None
        self._lock = threading.Lock()
        self._series: dict[str, list[list[float]]] = {}
        self._load()

    def _load(self) -> None:
        if self._path is None or not self._path.exists():
            return
        try:
            raw = json.loads(self._path.read_text(encoding="utf-8"))
            cutoff = time.time() - _THP_MAX_AGE_S
            self._series = {
                tech: [[float(t), float(dl), float(ul)] for t, dl, ul in pts if t >= cutoff]
                for tech, pts in (raw.get("series") or {}).items()
            }
            _log.debug(
                "throughput history loaded from %s: %s",
                self._path, {k: len(v) for k, v in self._series.items()},
            )
        except (OSError, ValueError, TypeError) as e:
            _log.warning("throughput history %s unreadable (%s) — starting empty", self._path, e)
            self._series = {}

    def _flush(self) -> None:
        if self._path is None:
            return
        tmp = self._path.with_suffix(self._path.suffix + ".tmp")
        try:
            tmp.parent.mkdir(parents=True, exist_ok=True)
            tmp.write_text(json.dumps({"series": self._series}), encoding="utf-8")
            tmp.replace(self._path)
        except OSError as e:
            _log.warning("throughput history %s not written: %s", self._path, e)

    def record(self, tech: str, dl: float, ul: float) -> None:
        now = time.time()
        with self._lock:
            pts = self._series.setdefault(tech, [])
            if pts and now - pts[-1][0] < _THP_STEP_S:
                return
            pts.append([now, round(dl, 1), round(ul, 1)])
            cutoff = now - _THP_MAX_AGE_S
            while pts and pts[0][0] < cutoff:
                pts.pop(0)
            self._flush()

    def record_network(self, network: Any) -> None:
        totals: dict[str, list[float]] = {}
        for cell in network.cells.values():
            t = totals.setdefault(_tech_of(cell), [0.0, 0.0])
            t[0] += (cell.thp_dl_kbps or 0.0) / 1000.0
            t[1] += (cell.thp_ul_kbps or 0.0) / 1000.0
        for tech, (dl, ul) in totals.items():
            self.record(tech, dl, ul)

    def window(self, tech: str, range_key: str | None) -> list[dict[str, Any]]:
        span = _TS_RANGES.get(range_key or "", _TS_RANGES[_TS_DEFAULT])
        floor = time.time() - span
        with self._lock:
            pts = [p for p in self._series.get(tech, []) if p[0] >= floor]
        return _resample(pts, span)


_ALERT_KIND = {"collision": "PCI Collision", "confusion": "PCI Confusion", "mod3": "Mod-3"}
_SEV_RANK = {"critical": 0, "major": 1, "minor": 2, "info": 3}


def _build_alerts(
    cells: list[dict[str, Any]],
    conflicts: list[dict[str, Any]],
    first_seen: dict[str, str],
    prb_warning: float = 80.0,
) -> list[dict[str, Any]]:
    now = _utcnow_iso()
    out: list[dict[str, Any]] = []

    def add(key: str, sev: str, cell: str, kind: str, msg: str) -> None:
        digest = hashlib.md5(key.encode("utf-8"), usedforsecurity=False).hexdigest()[:8]
        aid = f"AL-{digest}"
        iso = first_seen.setdefault(aid, now)
        out.append({
            "id": aid, "sev": sev, "t": iso[11:19], "iso": iso,
            "cell": cell, "kind": kind, "msg": msg,
        })

    for cf in conflicts:
        ids = cf.get("cells") or []
        add("|".join([cf["type"], str(cf.get("pci")), *sorted(map(str, ids))]),
            cf["severity"], ids[0] if ids else "—",
            _ALERT_KIND.get(cf["type"], "PCI"), cf["impact"])

    over = sorted((c for c in cells if c["prb"] >= prb_warning), key=lambda c: -c["prb"])
    congested = over[:10]
    if len(over) > len(congested):
        add(f"prb-summary|{prb_warning:.0f}", "major", "—", "PRB Congestion",
            f"{len(over)} cells above {prb_warning:.0f}% PRB utilisation — the {len(congested)} "
            f"worst are listed; raise the threshold in Settings to narrow")
    for c in congested:
        add(f"prb|{c['id']}", "major" if c["prb"] >= prb_warning + 10 else "minor", c["id"],
            "PRB Congestion",
            f"PRB utilisation at {c['prb']:.0f}% on {c['id']} (threshold {prb_warning:.0f}%)")

    out.sort(key=lambda a: (_SEV_RANK.get(a["sev"], 9), a["t"], a["cell"]))
    return out


def _tech_of(cell: Any) -> str:
    return str(getattr(cell.technology, "value", cell.technology)).lower()


def _cells_by_tech(network: Any) -> dict[str, int]:
    counts = {"lte": 0, "nr": 0}
    for cell in network.cells.values():
        t = _tech_of(cell)
        counts[t] = counts.get(t, 0) + 1
    return counts


def _resolve_tech(network: Any, tech: str | None) -> str:
    if tech in ("lte", "nr"):
        return tech
    counts = _cells_by_tech(network)
    return "nr" if counts["nr"] or not counts["lte"] else "lte"


def build_real_pci_data(
    network: Any,
    conflicts_result: dict[str, Any] | None,
    *,
    synthesize_coords: Any,
    alert_first_seen: dict[str, str] | None = None,
    tech: str | None = None,
    prb_warning: float = 80.0,
) -> dict[str, Any]:
    if network is None:
        return _empty_pci_data(mode="live", note="no PM data ingested yet")

    cells_by_tech = _cells_by_tech(network)
    tech = _resolve_tech(network, tech)

    cells_out: list[dict[str, Any]] = []
    for cell in network.cells.values():
        if _tech_of(cell) != tech:
            continue
        prb = round((cell.prb_util or 0.0) * 100, 0)
        real_pos = cell.lat is not None and cell.lon is not None
        if real_pos:
            lat, lng = cell.lat, cell.lon
        else:
            lat, lng = synthesize_coords(cell.id)
        neighbor_pcis: list[int] = []
        for n in network.neighbors_of(cell.id, same_tech_only=True):
            if n.pci not in neighbor_pcis:
                neighbor_pcis.append(n.pci)
        freq = cell.primary_frequency()
        mod_conflicts: dict[str, list[str]] = {"mod3": [], "mod4": [], "mod30": []}
        for n in network.neighbors_of(cell.id, same_tech_only=True):
            if freq is None or n.primary_frequency() != freq:
                continue
            if n.pci == cell.pci:
                continue
            for mod in (3, 4, 30):
                if n.pci % mod == cell.pci % mod:
                    mod_conflicts[f"mod{mod}"].append(n.id)
        if cell.band is not None:
            band_label = f"n{cell.band}" if tech == "nr" else f"B{cell.band}"
        elif freq is None:
            band_label = "—"
        elif tech == "lte":
            band_label = f"B{freq}"
        else:
            band_label = f"n{freq}"
        cells_out.append({
            "id": cell.id,
            "tech": tech,
            "site": cell.id.rsplit("-", 1)[0] if "-" in cell.id else cell.id,
            "region": cell.region or _UNASSIGNED_REGION,
            "pci": cell.pci,
            "band": band_label,
            "arfcn": freq,
            "sector": 1,
            "status": "active",
            "dl": round((cell.thp_dl_kbps or 0.0) / 1000.0, 1),
            "ul": round((cell.thp_ul_kbps or 0.0) / 1000.0, 1),
            "prb": prb,
            "ue": int(cell.active_ues or 0),
            "bler": round(cell.bler_dl or 0.0, 2),
            "cqi": round(cell.cqi or 0.0, 1),
            "sinr": round(cell.sinr_db or 0.0, 1),
            "lat": lat, "lng": lng, "approxPos": not real_pos,
            "neighbors": neighbor_pcis[:5],
            "neighborCount": len(neighbor_pcis),
            "modConflicts": mod_conflicts,
            "height_m": cell.antenna_height,
            "azimuth": cell.azimuth,
            "mech_tilt": None, "elec_tilt": None,
            "antenna_model": None, "antenna_gain_dbi": None,
            "beamwidth_deg": None, "tx_power_dbm": None,
        })

    ue_by_cell = {c["id"]: int(c["ue"]) for c in cells_out}
    conflicts_out: list[dict[str, Any]] = []
    if conflicts_result:
        for i, row in enumerate(conflicts_result.get("items", [])):
            cls = row.get("conflict_class", "")
            if cls == "collision":
                ctype = "collision"
            elif cls == "confusion":
                ctype = "confusion"
            else:
                ctype = "mod3"
            conflicts_out.append({
                "id": f"CFL-{2040 + i + 1}",
                "type": ctype,
                "pci": row.get("pci_a"),
                "severity": row.get("severity", "minor"),
                "cells": [row.get("cell_a_id"), row.get("cell_b_id")],
                "detected": _utcnow_iso(),
                "impact": (
                    f"{row.get('type_label', 'Conflict')} on "
                    f"{row.get('cell_a_id')} / {row.get('cell_b_id')} "
                    f"(PCI {row.get('pci_a')})"
                ),
                "affectedUe": ue_by_cell.get(str(row.get("cell_a_id")), 0)
                + ue_by_cell.get(str(row.get("cell_b_id")), 0),
            })

    total = len(cells_out)
    kpis = {
        "activeCells": total,
        "totalCells": total,
        "pciConflicts": len(conflicts_out),
        "pciTotal": len(conflicts_out),
        "coverage": 100.0 if total else 0.0,
        "coverageDelta": 0.0,
        "errorRate": round(sum(c["bler"] for c in cells_out) / total, 2) if total else 0.0,
        "errorRateDelta": 0.0,
        "connectedUe": sum(c["ue"] for c in cells_out),
        "connectedUeDelta": 0,
        "pulse": compute_pulse(cells_out, conflicts_out),
    }

    by_region: dict[str, list[dict[str, Any]]] = {}
    for c in cells_out:
        by_region.setdefault(c["region"], []).append(c)
    conflicted_ids = {cid for cf in conflicts_out for cid in cf["cells"]}

    regions_health = []
    for r in sorted(by_region):
        in_region = by_region[r]
        ids = {ic["id"] for ic in in_region}
        ok = sum(1 for ic in in_region if ic["id"] not in conflicted_ids)
        regions_health.append({
            "name": r,
            "cells": len(in_region),
            "ok": ok,
            "pct": round(100 * ok / len(in_region)) if in_region else 0,
            "conflicts": sum(1 for c in conflicts_out if ids & set(c["cells"])),
        })

    return {
        "CELLS": cells_out,
        "CONFLICTS": conflicts_out,
        "ALERTS": _build_alerts(
            cells_out, conflicts_out,
            alert_first_seen if alert_first_seen is not None else {},
            prb_warning=prb_warning,
        ),
        "TS": [],
        "SLICES": [],
        "KPIS": kpis,
        "REGIONS_HEALTH": regions_health,
        "meta": {
            "mode": "live",
            "updated": _utcnow_iso(),
            "poll_error": None,
            "tech": tech,
            "cellsByTech": cells_by_tech,
        },
    }


def _empty_pci_data(*, mode: str, note: str | None = None) -> dict[str, Any]:
    return {
        "CELLS": [], "CONFLICTS": [], "ALERTS": [],
        "TS": [], "HISTORY": [],
        "SLICES": [],
        "KPIS": {
            "activeCells": 0, "totalCells": 0, "pciConflicts": 0, "pciTotal": 0,
            "coverage": 0, "coverageDelta": 0, "errorRate": 0, "errorRateDelta": 0,
            "connectedUe": 0, "connectedUeDelta": 0,
            "pulse": compute_pulse([], []),
        },
        "REGIONS_HEALTH": [],
        "meta": {"mode": mode, "updated": _utcnow_iso(), "poll_error": note,
                 "tech": None, "cellsByTech": {"lte": 0, "nr": 0}},
    }


_DEFAULT_SETTINGS: dict[str, str] = {
    "account_name": "Operator",
    "account_timezone": "Asia/Kolkata",
    "theme_mode": "light",
    "pci_collision_threshold": "0",
    "pci_confusion_threshold": "1",
    "bler_threshold": "3.0",
    "prb_warning": "80",
    "mod3_detection": "warn_only",
    "brand_product_name": "PCI Planning and Optimization · rApp",
    "brand_logo_url": "",
    "brand_footer": "© 2026 Coran Labs",
    "brand_support_email": "support@coranlabs.com",
    "audit_retention_days": "90",
    "audit_immutable": "off",
    "notif_desktop": "on",
    "notif_sound_critical": "on",
    "notif_sound_minor": "off",
}

_EXCEL_HEADERS = [
    "cell_id", "lat", "lng",
    "site_name", "height_m", "azimuth",
    "mech_tilt", "elec_tilt",
    "antenna_model", "antenna_gain_dbi", "beamwidth_deg", "tx_power_dbm",
]


class DashboardData:

    def __init__(
        self,
        *,
        config: Any,
        network_cache: Any,
        services_module: Any,
    ) -> None:
        self._config = config
        self._network_cache = network_cache
        self._services = services_module
        self._lock = threading.RLock()

        self._settings: dict[str, str] = dict(_DEFAULT_SETTINGS)
        self._audit: list[dict[str, Any]] = []
        self._uploads: dict[int, dict[str, Any]] = {}
        self._next_upload_id = 1
        self._committed_pci: dict[str, tuple] = {}
        self._site_overrides: dict[str, dict[str, Any]] = {}
        self._alert_first_seen: dict[str, str] = {}
        self._thp = ThroughputHistory(getattr(config, "history_file", "") or "")
        self._history: dict[str, list[dict[str, Any]]] = {}

        self._log_audit(
            "dashboard_start", "info",
            "PCI dashboard started",
            source="system",
        )



    def snapshot(
        self, tech: str | None = None, ts_range: str | None = None,
    ) -> dict[str, Any]:
        cache = self._network_cache
        if cache is None:
            return _empty_pci_data(mode="live", note="config not loaded")
        snap = cache.get()
        network = self._live_network()
        conflicts_result = None
        if network is not None:
            tech = _resolve_tech(network, tech)
            try:
                conflicts_result = self._services.list_conflicts(
                    network, self._config, page_size=10_000, technology=tech,
                )
            except Exception as e:
                conflicts_result = None
                _ = e
        try:
            prb_warning = float(self._settings.get("prb_warning") or 80)
        except ValueError:
            prb_warning = 80.0
        prb_warning = min(100.0, max(1.0, prb_warning))
        data = build_real_pci_data(
            network, conflicts_result,
            synthesize_coords=self._services.synthesize_coords,
            alert_first_seen=self._alert_first_seen,
            tech=tech, prb_warning=prb_warning,
        )
        if snap.last_error:
            data["meta"]["poll_error"] = snap.last_error

        store = getattr(self._network_cache, "pm_store", None)
        age = getattr(store, "age_seconds", None) if store is not None else None
        data["meta"]["feed_age_seconds"] = round(age, 1) if age is not None else None
        data["HISTORY"] = self._record_history(data)
        if network is not None:
            self._thp.record_network(network)
        data["TS"] = self._thp.window(data["meta"]["tech"], ts_range)
        data["meta"]["ts_range"] = ts_range if ts_range in _TS_RANGES else _TS_DEFAULT
        return data

    def _record_history(self, data: dict[str, Any]) -> list[dict[str, Any]]:
        tech = data["meta"].get("tech") or "all"
        with self._lock:
            pts = self._history.setdefault(tech, [])
            now = time.time()
            if data["CELLS"] and (not pts or now - pts[-1]["ts"] >= _HISTORY_STEP_S):
                conflicts = data["CONFLICTS"]
                involved = {cid for cf in conflicts for cid in cf["cells"]}
                pts.append({
                    "ts": now,
                    "iso": datetime.fromtimestamp(now, UTC).isoformat(),
                    "collision": sum(1 for c in conflicts if c["type"] == "collision"),
                    "confusion": sum(1 for c in conflicts if c["type"] == "confusion"),
                    "modn": sum(1 for c in conflicts if c["type"] == "mod3"),
                    "cells": len(data["CELLS"]),
                    "conflictCells": len(involved),
                })
                del pts[:-_HISTORY_MAX]
            return list(pts)


    def replan_propose(self, cell_id: str) -> dict[str, Any]:
        return self._live_propose(cell_id)

    def replan_commit(self, cell_id: str, proposed_pci: Any) -> dict[str, Any]:
        return self._live_commit(cell_id, proposed_pci)


    def _live_network(self) -> Any:
        cache = self._network_cache
        if cache is None:
            return None
        network = cache.get().network
        if network is not None:
            self._reapply_overrides(network)
        return network

    def _reapply_overrides(self, network: Any) -> None:
        if getattr(network, _APPLIED_MARKER, False):
            return
        now = datetime.now(UTC).timestamp()
        with self._lock:
            for cell_id, (pci, committed_at) in list(self._committed_pci.items()):
                cell = network.cells.get(cell_id)
                if cell is None:
                    del self._committed_pci[cell_id]
                    continue
                if cell.pci == pci:
                    del self._committed_pci[cell_id]
                    continue
                if now - committed_at > _COMMIT_CONFIRM_WINDOW_S:
                    del self._committed_pci[cell_id]
                    self._log_audit(
                        "pci_replan_unconfirmed", "critical",
                        f"{cell_id}: PCI {pci} was committed to SDNR but the PM feed "
                        f"still reports {cell.pci} after "
                        f"{int(_COMMIT_CONFIRM_WINDOW_S // 60)} minutes — showing the "
                        f"feed's value again",
                        cell_id=cell_id, pci_old=cell.pci, pci_new=pci, source="system",
                    )
                    continue
                cell.pci = pci
                if cell.pci_components is not None:
                    cell.pci_components = (pci // 3, pci % 3)

            for cell_id, fields in self._site_overrides.items():
                cell = network.cells.get(cell_id)
                if cell is None:
                    continue
                for name, value in fields.items():
                    setattr(cell, name, value)
        setattr(network, _APPLIED_MARKER, True)

    def _live_propose(self, cell_id: str) -> dict[str, Any]:
        network = self._live_network()
        if network is None:
            return {"ok": False, "error": "No network snapshot loaded"}
        cell = network.cells.get(cell_id)
        if cell is None:
            return {"ok": False, "error": f"Cell {cell_id} not found"}

        neighbours = network.neighbors_of(cell_id, same_tech_only=True)
        freq = cell.primary_frequency()
        same_freq = [n for n in neighbours if n.primary_frequency() == freq]
        blocked = {n.pci for n in same_freq}
        blocked_mod3 = {n.pci % 3 for n in same_freq}

        lo, hi = _pci_pool_range(self._config, cell)
        clean = next(
            (p for p in range(lo, hi)
             if p != cell.pci and p not in blocked and (p % 3) not in blocked_mod3),
            None,
        )
        proposed = clean if clean is not None else next(
            (p for p in range(lo, hi) if p != cell.pci and p not in blocked),
            None,
        )
        if proposed is None:
            return {"ok": False, "error": "No free PCI in the pool"}

        used_in_pool = {
            c.pci for c in network.cells.values()
            if c.primary_frequency() == freq and lo <= c.pci < hi
        }
        reasons = [
            f"clear of {len(same_freq)} co-frequency neighbour"
            f"{'' if len(same_freq) == 1 else 's'}",
            f"inside the {cell.cell_type} pool {lo}–{hi - 1}",
        ]
        if clean is not None:
            reasons.append(f"mod-3 residue {proposed % 3} unused by neighbours")
        return {
            "ok": True,
            "cell_id": cell_id,
            "old_pci": cell.pci,
            "proposed_pci": proposed,
            "neighbors_checked": len(same_freq),
            "free_pcis_remaining": (hi - lo) - len(used_in_pool),
            "mod3_safe": clean is not None,
            "reason": " · ".join(reasons),
        }

    def _sdnr_client_for(self, cell: Any) -> Any:
        from pci_planning_and_optimization.sdnr import make_sdnr_client
        from pci_planning_and_optimization.sdnr.client import _managed_element

        client = make_sdnr_client(self._config)
        if client is not None:
            return client

        sdnr_cfg = getattr(self._config, "sdnr", None)
        if not (getattr(sdnr_cfg, "enabled", False) and getattr(sdnr_cfg, "base_url", "")):
            return None
        mount = _managed_element(cell.dn or "", default="")
        if not mount:
            return None

        from pci_planning_and_optimization.sdnr.client import SdnrClient
        return SdnrClient(
            base_url=sdnr_cfg.base_url,
            username=sdnr_cfg.username,
            password=sdnr_cfg.password,
            netconf_node_id=mount,
            function_id=sdnr_cfg.function_id,
            timeout_s=sdnr_cfg.timeout_s,
        )

    def _live_commit(self, cell_id: str, proposed_pci: Any) -> dict[str, Any]:
        from pci_planning_and_optimization.sdnr import apply_change, preflight

        network = self._live_network()
        if network is None:
            return {"ok": False, "error": "No network snapshot loaded"}
        cell = network.cells.get(cell_id)
        if cell is None:
            return {"ok": False, "error": f"Cell {cell_id} not found"}
        try:
            new_pci = int(proposed_pci)
        except (TypeError, ValueError):
            return {"ok": False, "error": f"proposed_pci is not an integer: {proposed_pci!r}"}

        client = self._sdnr_client_for(cell)
        if client is None:
            sdnr_cfg = getattr(self._config, "sdnr", None)
            if not getattr(sdnr_cfg, "enabled", False):
                err = ("SDNR writes are disabled (sdnr.enabled=false). "
                       "Set sdnr.enabled=true to commit PCI changes to the network.")
            elif not getattr(sdnr_cfg, "base_url", ""):
                err = "sdnr.base_url is empty — cannot reach SDNR"
            else:
                err = (f"no NETCONF mount for {cell_id}: sdnr.netconf_node_id is empty and "
                       "the cell reported no ManagedElement to fall back on")
            self._log_audit(
                "pci_replan_failed", "major",
                f"PCI re-plan {cell_id} {cell.pci} → {new_pci} not sent: {err}",
                cell_id=cell_id, pci_old=cell.pci, pci_new=new_pci, source="operator",
            )
            return {"ok": False, "error": err}

        ok, note = preflight(client)
        if not ok:
            self._log_audit(
                "pci_replan_failed", "critical",
                f"SDNR preflight failed for {cell_id}: {note}",
                cell_id=cell_id, pci_old=cell.pci, pci_new=new_pci, source="operator",
            )
            return {"ok": False, "error": f"SDNR preflight failed: {note}"}

        old_pci = cell.pci
        ok, note = apply_change(_change_record(cell, new_pci), client)
        self._log_audit(
            "pci_replan" if ok else "pci_replan_failed",
            "info" if ok else "critical",
            (f"PCI change committed to SDNR: {cell_id} {old_pci} → {new_pci}"
             if ok else f"SDNR write failed for {cell_id}: {note}"),
            cell_id=cell_id, pci_old=old_pci, pci_new=new_pci, source="operator",
        )
        if not ok:
            return {"ok": False, "error": note}

        with self._lock:
            self._committed_pci[cell_id] = (
                new_pci, datetime.now(UTC).timestamp(),
            )
        try:
            delattr(network, _APPLIED_MARKER)
        except AttributeError:
            pass
        self._reapply_overrides(network)
        return {
            "ok": True,
            "cell_id": cell_id,
            "old_pci": old_pci,
            "new_pci": new_pci,
            "sdnr_note": note,
            "audit_msg": f"PCI re-plan: {cell_id} {old_pci} → {new_pci} (operator commit, SDNR)",
        }


    def _log_audit(
        self,
        event_type: str,
        severity: str,
        description: str,
        *,
        cell_id: str | None = None,
        pci_old: int | None = None,
        pci_new: int | None = None,
        source: str = "system",
    ) -> None:
        with self._lock:
            actor = None if source == "system" else self._actor()
            self._audit.append({
                "id": len(self._audit) + 1,
                "timestamp": _utcnow_iso(),
                "event_type": event_type,
                "severity": severity,
                "cell_id": cell_id,
                "pci_old": pci_old,
                "pci_new": pci_new,
                "description": description,
                "source": source,
                "actor": actor,
            })

    def _actor(self) -> dict[str, str]:
        from pci_planning_and_optimization.api.auth import current_user

        return {
            "name": current_user.get("")
            or self._settings.get("account_name")
            or "Operator",
        }

    def add_audit(self, body: dict[str, Any]) -> None:
        event_type = str(body.get("event_type") or "manual")
        description = str(body.get("description") or "")
        if event_type in ("sign_in", "sign_out"):
            who = self._actor()
            verb = "signed in" if event_type == "sign_in" else "signed out"
            description = f"{who['name']} {verb}"
        self._log_audit(
            event_type,
            str(body.get("severity") or "info"),
            description,
            cell_id=body.get("cell_id"),
            pci_old=body.get("pci_old"),
            pci_new=body.get("pci_new"),
            source="operator",
        )

    def get_audit(
        self,
        *,
        page: int = 1,
        per_page: int = 50,
        severity: str | None = None,
        event_type: str | None = None,
        search: str | None = None,
    ) -> dict[str, Any]:
        with self._lock:
            logs = list(reversed(self._audit))

        if severity:
            logs = [e for e in logs if e.get("severity") == severity]
        if event_type:
            logs = [e for e in logs if e.get("event_type") == event_type]
        if search:
            s = search.lower()
            logs = [
                e for e in logs
                if s in (e.get("description") or "").lower()
                or s in (e.get("cell_id") or "").lower()
                or s in ((e.get("actor") or {}).get("name") or "").lower()
            ]

        total = len(logs)
        start = max(0, (page - 1) * per_page)
        return {
            "total": total,
            "page": page,
            "per_page": per_page,
            "logs": logs[start:start + per_page],
        }

    def audit_stats(self) -> dict[str, Any]:
        with self._lock:
            logs = list(self._audit)
        by_severity: dict[str, int] = {}
        by_type: dict[str, int] = {}
        for entry in logs:
            sv = entry.get("severity", "info")
            by_severity[sv] = by_severity.get(sv, 0) + 1
            et = entry.get("event_type", "event")
            by_type[et] = by_type.get(et, 0) + 1
        return {"total": len(logs), "by_severity": by_severity, "by_type": by_type}


    def get_settings(self) -> dict[str, str]:
        with self._lock:
            return dict(self._settings)

    def set_settings(self, body: dict[str, Any]) -> None:
        with self._lock:
            for k, v in body.items():
                self._settings[str(k)] = str(v)
        self._log_audit(
            "settings_change", "info",
            f"Settings updated: {', '.join(sorted(body.keys()))}",
            source="operator",
        )

    def excel_template_bytes(self, tech: str | None = None) -> bytes:
        import openpyxl

        tech_norm = (tech or "").lower()
        if tech_norm in ("lte", "4g"):
            source_cells: list[dict[str, Any]] = list(self.snapshot("lte").get("CELLS", []))
            sheet_title = "LTE Cell Plan"
        elif tech_norm in ("nr", "5g"):
            source_cells = list(self.snapshot("nr").get("CELLS", []))
            sheet_title = "5G NR Cell Plan"
        else:
            source_cells = list(self.snapshot().get("CELLS", []))
            sheet_title = "Site & Antenna Config"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        ws.append(_EXCEL_HEADERS)
        col_widths = [22, 12, 12, 22, 11, 10, 11, 11, 22, 18, 15, 14]
        for i, w in enumerate(col_widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w
        bold = openpyxl.styles.Font(bold=True)
        fill_req = openpyxl.styles.PatternFill("solid", fgColor="FFFCE4D6")
        required_cols = {0, 1, 2, 5}
        for idx, cell in enumerate(ws[1]):
            cell.font = bold
            if idx in required_cols:
                cell.fill = fill_req

        for i, c in enumerate(source_cells):
            ws.append([
                c.get("id"), c.get("lat"), c.get("lng"),
                c.get("site"), c.get("height_m"), _default_azimuth(c, i),
                c.get("mech_tilt"), c.get("elec_tilt"),
                c.get("antenna_model"), c.get("antenna_gain_dbi"),
                c.get("beamwidth_deg"), c.get("tx_power_dbm"),
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def full_pool_bytes(self, tech: str | None = None) -> bytes:
        import openpyxl

        tech_norm = "lte" if (tech or "").lower() in ("lte", "4g") else "nr"
        pool_size = 504 if tech_norm == "lte" else 1008
        snap = self.snapshot(tech_norm)
        cells = list(snap.get("CELLS", []))
        mod3_cell_ids = {
            cid for cf in snap.get("CONFLICTS", [])
            if cf.get("type") == "mod3"
            for cid in (cf.get("cells") or [])
        }
        sheet_title = f"{'LTE' if tech_norm == 'lte' else '5G NR'} PCI Pool (0..{pool_size - 1})"

        by_pci: dict[int, list[dict[str, Any]]] = {}
        for c in cells:
            try:
                p = int(c.get("pci"))
            except (TypeError, ValueError):
                continue
            if 0 <= p < pool_size:
                by_pci.setdefault(p, []).append(c)

        mod3_pcis: set = set()
        for c in cells:
            if c.get("id") in mod3_cell_ids:
                try:
                    mod3_pcis.add(int(c["pci"]))
                except (TypeError, ValueError):
                    pass

        headers = [
            "pci", "status", "cell_count", "assigned_cells",
            "regions", "bands", "arfcns",
            "mod3", "mod6", "mod30", "notes",
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        ws.append(headers)
        col_widths = [6, 14, 7, 60, 28, 18, 18, 6, 6, 7, 50]
        for i, w in enumerate(col_widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w
        bold = openpyxl.styles.Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
        ws.freeze_panes = "A2"

        for pci in range(pool_size):
            holders = by_pci.get(pci, [])
            if not holders:
                status = "unassigned"
            elif len(holders) > 1:
                status = "collision"
            elif pci in mod3_pcis:
                status = "mod3_cluster"
            else:
                status = "in_use"

            ids = [c.get("id", "") for c in holders]
            regions = sorted({c.get("region") for c in holders if c.get("region")})
            bands = sorted({str(c.get("band")) for c in holders if c.get("band")})
            arfcns = sorted({c.get("arfcn") for c in holders if c.get("arfcn")})

            note = ""
            if status == "collision":
                shared_arfcns = ", ".join(str(a) for a in arfcns)
                note = f"PCI {pci} shared by {len(holders)} cells on ARFCN(s) {shared_arfcns}"
            elif status == "mod3_cluster":
                note = f"part of a mod-3 group {pci % 3} cluster (reference-signal interference)"
            elif status == "unassigned":
                note = "free PCI slot — available for re-plan"

            ws.append([
                pci,
                status,
                len(holders),
                "; ".join(ids),
                ", ".join(regions),
                ", ".join(bands),
                ", ".join(str(a) for a in arfcns),
                pci % 3,
                pci % 6,
                pci % 30,
                note,
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def excel_upload(self, filename: str, raw: bytes, description: str) -> dict[str, Any]:
        import openpyxl

        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active
            headers = [
                str(c.value).strip().lower() if c.value else f"col{i}"
                for i, c in enumerate(ws[1])
            ]
            records: list[dict[str, Any]] = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(v is None for v in row):
                    continue
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                rec = _parse_excel_row(row_idx, row_dict)
                records.append(rec)
        except Exception as e:
            return {"error": f"Failed to parse file: {e}"}

        if not records:
            return {"error": "No data rows found in the file."}

        bad = [r for r in records if not r["valid"]]
        if bad:
            return {
                "ok": False,
                "error": (
                    f"Upload rejected — {len(bad)} of {len(records)} rows have "
                    f"validation errors. Required columns: cell_id, lat, lng, azimuth."
                ),
                "invalid_rows": [
                    {"row": r["row_num"], "cell_id": r["cell_id"] or None,
                     "reason": r["error_msg"]}
                    for r in bad[:50]
                ],
                "invalid_count": len(bad),
                "total_rows": len(records),
            }

        with self._lock:
            upload_id = self._next_upload_id
            self._next_upload_id += 1
            self._uploads[upload_id] = {
                "id": upload_id,
                "filename": filename,
                "upload_time": _utcnow_iso(),
                "description": description,
                "row_count": len(records),
                "applied_at": None,
                "applied_by": None,
                "applied_count": 0,
                "skipped_count": 0,
                "records": records,
            }
        self._log_audit(
            "excel_upload", "info",
            f"Site config uploaded: {filename} ({len(records)} rows)",
            source="operator",
        )
        return {"ok": True, "upload_id": upload_id, "rows": len(records)}

    def excel_uploads(self) -> list[dict[str, Any]]:
        with self._lock:
            ups = sorted(self._uploads.values(),
                         key=lambda u: u["upload_time"], reverse=True)
            return [{
                "id": u["id"],
                "filename": u["filename"],
                "upload_time": u["upload_time"],
                "description": u["description"],
                "row_count": u["row_count"],
                "applied_at": u["applied_at"],
                "applied_by": u["applied_by"],
                "applied_count": u["applied_count"],
                "skipped_count": u["skipped_count"],
                "status": "applied" if u["applied_at"] else "pending",
            } for u in ups]

    def excel_records(
        self, upload_id: int, *, page: int = 1, per_page: int = 50,
        errors_only: bool = False,
    ) -> dict[str, Any] | None:
        with self._lock:
            up = self._uploads.get(upload_id)
            if up is None:
                return None
            recs = up["records"]
        if errors_only:
            recs = [r for r in recs if not r["valid"]]
        total = len(recs)
        start = max(0, (page - 1) * per_page)
        return {
            "total": total, "page": page, "per_page": per_page,
            "records": recs[start:start + per_page],
        }

    def excel_delete(self, upload_id: int) -> bool:
        with self._lock:
            if upload_id in self._uploads:
                del self._uploads[upload_id]
                return True
            return False

    _SITE_FIELD_MAP = {
        "lat": "lat", "lng": "lon", "azimuth": "azimuth", "height_m": "antenna_height",
    }

    def _live_excel_apply(self, up: dict[str, Any]) -> dict[str, Any]:
        network = self._live_network()
        if network is None:
            return {"ok": False, "error": "No network snapshot loaded"}

        applied: list[dict[str, Any]] = []
        no_op: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        unstored: set = set()
        total_field_changes = 0

        for r in up["records"]:
            cid = (r.get("cell_id") or "").strip()
            if not r["valid"]:
                skipped.append({"row": r["row_num"], "cell_id": cid,
                                "reason": r["error_msg"] or "invalid row"})
                continue
            if not cid:
                skipped.append({"row": r["row_num"], "reason": "empty cell_id"})
                continue
            cell = network.cells.get(cid)
            if cell is None:
                skipped.append({"row": r["row_num"], "cell_id": cid,
                                "reason": "cell not in live network"})
                continue

            for sheet_name in ("mech_tilt", "elec_tilt", "antenna_model",
                               "antenna_gain_dbi", "beamwidth_deg", "tx_power_dbm"):
                if r.get(sheet_name) is not None:
                    unstored.add(sheet_name)

            changed = {}
            for sheet_name, model_name in self._SITE_FIELD_MAP.items():
                value = r.get(sheet_name)
                if value is None:
                    continue
                if getattr(cell, model_name, None) == value:
                    continue
                changed[model_name] = value

            if not changed:
                no_op.append({"row": r["row_num"], "cell_id": cid,
                              "reason": "all fields already match live cell"})
                continue

            with self._lock:
                self._site_overrides.setdefault(cid, {}).update(changed)
            for model_name, value in changed.items():
                setattr(cell, model_name, value)
            total_field_changes += len(changed)
            applied.append({"row": r["row_num"], "cell_id": cid,
                            "changed": list(changed), "change_count": len(changed)})

        with self._lock:
            up["applied_at"] = _utcnow_iso()
            up["applied_by"] = self._settings.get("account_name", "operator")
            up["applied_count"] = len(applied)
            up["skipped_count"] = len(skipped) + len(no_op)
            filename = up["filename"]

        note = (f'Site config "{filename}" applied: {len(applied)} cells updated, '
                f"{total_field_changes} field changes, {len(no_op)} no-op, "
                f"{len(skipped)} skipped")
        if unstored:
            note += f" ({', '.join(sorted(unstored))} not stored — no field on the cell model)"
        self._log_audit("site_config_apply", "info", note, source="operator")

        return {
            "ok": True,
            "applied": applied,
            "no_op": no_op,
            "skipped": skipped,
            "applied_count": len(applied),
            "field_changes": total_field_changes,
            "skipped_count": len(skipped) + len(no_op),
            "unstored_columns": sorted(unstored),
            "new_conflicts": [],
            "note": note,
        }

    def excel_apply(self, upload_id: int) -> dict[str, Any] | None:
        with self._lock:
            up = self._uploads.get(upload_id)
            if up is None:
                return None

        return self._live_excel_apply(up)


def _default_azimuth(cell: dict[str, Any], idx: int) -> int:
    az = cell.get("azimuth")
    if az is not None:
        return int(az)
    sector = cell.get("sector")
    if isinstance(sector, int) and 1 <= sector <= 6:
        return ((sector - 1) % 3) * 120 + (60 if sector > 3 else 0)
    return (idx % 3) * 120


def _parse_excel_row(row_idx: int, row_dict: dict[str, Any]) -> dict[str, Any]:
    def get(keys: list[str]) -> Any:
        for k in keys:
            v = row_dict.get(k)
            if v is not None and str(v).strip() != "":
                return v
        return None

    def safe_int(v: Any) -> int | None:
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None

    def safe_float(v: Any) -> float | None:
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def safe_str(v: Any) -> str | None:
        if v is None:
            return None
        s = str(v).strip()
        return s or None

    cell_id = safe_str(get(["cell_id", "cell id", "cellid", "managed_element", "managedelement"]))
    lat = safe_float(get(["lat", "latitude"]))
    lng = safe_float(get(["lng", "lon", "long", "longitude"]))
    site_name = safe_str(get(["site_name", "site name", "sitename", "site"]))
    height_m = safe_float(get(["height_m", "height", "height (m)", "antenna_height", "antenna height"]))
    azimuth = safe_int(get(["azimuth", "azimuth_deg", "az"]))
    mech_tilt = safe_float(get(["mech_tilt", "mechanical_tilt", "mechanical tilt", "mtilt"]))
    elec_tilt = safe_float(get(["elec_tilt", "electrical_tilt", "electrical tilt", "etilt", "ret_tilt"]))
    ant_model = safe_str(get(["antenna_model", "antenna model", "antenna"]))
    ant_gain = safe_float(get(["antenna_gain_dbi", "antenna gain", "gain_dbi", "gain"]))
    beamwidth = safe_int(get(["beamwidth_deg", "beamwidth", "horizontal_beamwidth", "hbw"]))
    tx_power = safe_float(get(["tx_power_dbm", "tx_power", "tx power", "power_dbm"]))

    errors: list[str] = []
    if not cell_id:
        errors.append("cell_id is required")
    if lat is None:
        errors.append("lat is required")
    elif not (-90.0 <= lat <= 90.0):
        errors.append(f"lat {lat} out of range -90..90")
    if lng is None:
        errors.append("lng is required")
    elif not (-180.0 <= lng <= 180.0):
        errors.append(f"lng {lng} out of range -180..180")
    if azimuth is None:
        errors.append("azimuth is required")
    elif not (0 <= azimuth <= 359):
        errors.append(f"azimuth {azimuth} out of range 0..359")
    if height_m is not None and not (0.0 <= height_m <= 200.0):
        errors.append(f"height_m {height_m} out of range 0..200")
    if mech_tilt is not None and not (-10.0 <= mech_tilt <= 30.0):
        errors.append(f"mech_tilt {mech_tilt} out of range -10..30")
    if elec_tilt is not None and not (-10.0 <= elec_tilt <= 30.0):
        errors.append(f"elec_tilt {elec_tilt} out of range -10..30")
    if ant_gain is not None and not (0.0 <= ant_gain <= 25.0):
        errors.append(f"antenna_gain_dbi {ant_gain} out of range 0..25")
    if beamwidth is not None and not (30 <= beamwidth <= 360):
        errors.append(f"beamwidth_deg {beamwidth} out of range 30..360")
    if tx_power is not None and not (0.0 <= tx_power <= 60.0):
        errors.append(f"tx_power_dbm {tx_power} out of range 0..60")

    return {
        "row_num": row_idx,
        "cell_id": cell_id or "",
        "lat": lat, "lng": lng,
        "site_name": site_name,
        "height_m": height_m,
        "azimuth": azimuth,
        "mech_tilt": mech_tilt,
        "elec_tilt": elec_tilt,
        "antenna_model": ant_model,
        "antenna_gain_dbi": ant_gain,
        "beamwidth_deg": beamwidth,
        "tx_power_dbm": tx_power,
        "valid": len(errors) == 0,
        "error_msg": "; ".join(errors) if errors else None,
        "extra_data": {},
    }
